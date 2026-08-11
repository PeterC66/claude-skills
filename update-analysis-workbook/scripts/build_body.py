"""Build the new LedgerDetails body from a Beacon ledger download.

Reads the current Analysis workbook (for its Categories rules and the NewCat
allocations already made) plus a fresh Beacon download, and writes:

  * a values-only .xlsx holding the rebuilt 9-column body of the LedgerDetails
    table, ready for apply_update.ps1 to paste in; and
  * a .json report of everything that changed, which is the thing a human
    actually needs to read before committing to the update.

Nothing is written to the Analysis workbook here - this step is read-only and
safe to run as often as you like.

Usage:
  python build_body.py --workbook <analysis.xlsm> --download <beacon.xlsx>
                       --out <body.xlsx> --report <report.json>
"""
import argparse
import json
import re
import sys
from collections import defaultdict

import openpyxl

LD_COLS = ["tkey", "category", "amount", "date", "account", "payee", "detail",
           "notes", "NewCat"]


def header_map(ws):
    """{lower-cased header: 0-based column index} from row 1."""
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(v).strip().lower(): i for i, v in enumerate(row) if v is not None}


def need(hm, name, sheet):
    key = name.lower()
    if key not in hm:
        sys.exit(f"ERROR: sheet '{sheet}' has no '{name}' column. Found: {sorted(hm)}")
    return hm[key]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True, help="the Analysis .xlsm as it stands now (or its backup)")
    ap.add_argument("--download", required=True, help="the Beacon ledger .xlsx")
    ap.add_argument("--out", required=True, help="body .xlsx to write")
    ap.add_argument("--report", required=True, help="report .json to write")
    args = ap.parse_args()

    # ---- the Analysis workbook as it stands ---------------------------------
    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    for s in ("Categories", "LedgerDetails"):
        if s not in wb.sheetnames:
            sys.exit(f"ERROR: '{args.workbook}' has no {s} sheet - is it the right workbook?")

    cats = [(r[0], r[1]) for r in wb["Categories"].iter_rows(min_row=2, values_only=True) if r[0]]

    ld = wb["LedgerDetails"]
    hm = header_map(ld)
    idx = [need(hm, c, "LedgerDetails") for c in LD_COLS]
    old_rows = []
    for r in ld.iter_rows(min_row=2, values_only=True):
        if r[idx[0]] is None:
            continue
        old_rows.append([r[i] if i < len(r) else None for i in idx])
    wb.close()

    def auto_newcat(category):
        """First Categories rule whose RegExp matches - mirrors FindFirstMatch in Refresh.bas."""
        for name, pattern in cats:
            if pattern and re.search(str(pattern), str(category), re.I):
                return name
        return ""

    # ---- the fresh download -------------------------------------------------
    wb = openpyxl.load_workbook(args.download, read_only=True, data_only=True)
    for s in ("Ledger", "Detail"):
        if s not in wb.sheetnames:
            sys.exit(f"ERROR: '{args.download}' has no {s} sheet - is it a Beacon ledger export?")

    lh = header_map(wb["Ledger"])
    li = {c: need(lh, c, "Ledger") for c in ("tkey", "date", "account", "payee", "detail", "notes")}
    ledger = {}
    for r in wb["Ledger"].iter_rows(min_row=2, values_only=True):
        if r[li["tkey"]] is None:
            continue
        ledger[r[li["tkey"]]] = tuple(r[li[c]] for c in ("date", "account", "payee", "detail", "notes"))

    dh = header_map(wb["Detail"])
    di = {c: need(dh, c, "Detail") for c in ("tkey", "category", "amount")}
    detail = [(r[di["tkey"]], r[di["category"]], r[di["amount"]])
              for r in wb["Detail"].iter_rows(min_row=2, values_only=True) if r[di["tkey"]] is not None]
    wb.close()

    # ---- pair old rows to new ones on (tkey, category) ----------------------
    # The pair is not guaranteed unique - one transaction can carry the same
    # category twice - so match the nth occurrence on each side.
    new_by_key = defaultdict(list)
    for i, (tkey, cat, _amt) in enumerate(detail):
        new_by_key[(tkey, str(cat))].append(i)

    used = set()
    out_rows = []
    report = {"workbook": args.workbook, "download": args.download,
              "updated": [], "dropped": [], "added": 0, "added_blank": 0,
              "blank_rows": [], "not_found": []}

    def enrich(tkey, amt):
        lr = ledger.get(tkey)
        if lr is None:
            nf = "Ledger NOT FOUND"
            report["not_found"].append(tkey)
            return [amt, None, nf, nf, nf, nf]
        return [amt, lr[0], lr[1], lr[2], lr[3], lr[4]]

    seen = defaultdict(int)
    for row in old_rows:
        key = (row[0], str(row[1]))
        n = seen[key]
        seen[key] += 1
        idxs = new_by_key.get(key, [])
        if n >= len(idxs):
            report["dropped"].append({"tkey": row[0], "category": row[1], "amount": row[2],
                                      "date": row[3], "payee": row[5], "newcat": row[8]})
            continue
        i = idxs[n]
        used.add(i)
        tkey, cat, amt = detail[i]
        vals = enrich(tkey, amt)
        changed = {}
        for label, before, after in zip(["amount", "date", "account", "payee", "detail", "notes"],
                                        [row[2], row[3], row[4], row[5], row[6], row[7]], vals):
            if label == "amount":
                same = before is not None and after is not None and abs(float(before) - float(after)) < 1e-9
            else:
                same = (before or "") == (after or "")
            if not same:
                changed[label] = [before, after]
        if changed:
            report["updated"].append({"tkey": tkey, "category": cat, "newcat": row[8], "changes": changed})
        out_rows.append([tkey, cat] + vals + [row[8]])

    for i, (tkey, cat, amt) in enumerate(detail):
        if i in used:
            continue
        vals = enrich(tkey, amt)
        nc = auto_newcat(cat)
        report["added"] += 1
        if not nc:
            report["added_blank"] += 1
            report["blank_rows"].append({"tkey": tkey, "category": cat, "amount": amt,
                                         "date": vals[1], "account": vals[2],
                                         "payee": vals[3], "detail": vals[4]})
        out_rows.append([tkey, cat] + vals + [nc])

    # ---- write the paste-ready body ----------------------------------------
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Body"
    for r in out_rows:
        # dates stay TEXT - the Year column coerces them, and letting Excel turn
        # them into real dates silently reinterprets dd/mm as mm/dd
        ws.append([r[0], r[1], r[2], (str(r[3]) if r[3] is not None else None),
                   r[4], r[5], r[6], r[7], r[8]])
    for cell in ws["D"]:
        cell.number_format = "@"
    out.save(args.out)

    report["rows_before"] = len(old_rows)
    report["rows_after"] = len(out_rows)
    with open(args.report, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=1, default=str)

    print(f"rows before : {len(old_rows)}")
    print(f"rows after  : {len(out_rows)}")
    print(f"appended    : {report['added']} ({report['added_blank']} need a NewCat)")
    print(f"dropped     : {len(report['dropped'])}")
    print(f"amended     : {len(report['updated'])} existing rows changed in Beacon")
    print(f"ledger n/f  : {len(report['not_found'])}")
    print(f"\nbody   -> {args.out}\nreport -> {args.report}")


if __name__ == "__main__":
    main()
