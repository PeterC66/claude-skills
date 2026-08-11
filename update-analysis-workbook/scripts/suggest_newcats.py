"""Work out what the unallocated rows should be, using the workbook's own history.

Every blank NewCat is a judgement call, but the workbook already contains
thousands of those judgements. For each unallocated row this prints the
transaction and how that same Beacon category has been allocated before, so the
allocation can be proposed from precedent rather than guessed - and so that
genuinely new Beacon categories (which need a rule, not a one-off allocation)
stand out from the ambiguous ones.

Run it AFTER apply_update.ps1, against the updated workbook.

Usage:
  python suggest_newcats.py --workbook <analysis.xlsm> [--json out.json]
"""
import argparse
import json
from collections import Counter, defaultdict

import openpyxl

COLS = ["tkey", "category", "amount", "date", "account", "payee", "detail", "notes", "NewCat"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--json", help="also write the findings as JSON")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    ws = wb["LedgerDetails"]
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ix = {str(h).strip().lower(): i for i, h in enumerate(hdr) if h}
    idx = [ix[c.lower()] for c in COLS]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[idx[0]] is None:
            continue
        rows.append([r[i] if i < len(r) else None for i in idx])
    cats = {r[0] for r in wb["Categories"].iter_rows(min_row=2, values_only=True) if r[0]}
    wb.close()

    # how each Beacon category has been allocated before, and by which payee
    hist = defaultdict(Counter)
    hist_payee = defaultdict(lambda: defaultdict(Counter))
    for r in rows:
        if r[8]:
            hist[r[1]][r[8]] += 1
            hist_payee[r[1]][str(r[5] or "")[:30]][r[8]] += 1

    blanks = [r for r in rows if not r[8]]
    if not blanks:
        print("No blank NewCat rows - nothing to allocate.")
        return

    by_cat = defaultdict(list)
    for r in blanks:
        by_cat[r[1]].append(r)

    unseen = {c: v for c, v in by_cat.items() if c not in hist}
    seen = {c: v for c, v in by_cat.items() if c in hist}

    out = {"blank_total": len(blanks), "new_categories": {}, "ambiguous": {}}

    print(f"{len(blanks)} unallocated row(s) in {len(by_cat)} Beacon categor(y/ies)\n")

    if unseen:
        print("=" * 70)
        print("BEACON CATEGORIES NEVER SEEN BEFORE")
        print("These need a RULE on the Categories sheet (column B regexp), not a")
        print("one-off allocation, or they will come back blank every update.")
        print("=" * 70)
        for cat, rs in sorted(unseen.items(), key=lambda x: -len(x[1])):
            total = sum(r[2] or 0 for r in rs)
            print(f"\n--- {cat!r}: {len(rs)} row(s), total {total:.2f}")
            for r in rs[:12]:
                print(f"    {r[3]}  {str(r[4]):<14} {r[2]:>9.2f}  {str(r[5])[:28]:<28} {str(r[6] or '')[:45]}")
            if len(rs) > 12:
                print(f"    ... and {len(rs) - 12} more")
            out["new_categories"][cat] = {"count": len(rs), "total": round(total, 2)}

    if seen:
        print("\n" + "=" * 70)
        print("CATEGORIES SEEN BEFORE BUT SPLIT ACROSS SEVERAL NEW CATEGORIES")
        print("These are per-transaction judgement calls. The history below is")
        print("the strongest guide - especially the same payee's history.")
        print("=" * 70)
        for cat, rs in sorted(seen.items(), key=lambda x: -len(x[1])):
            print(f"\n--- {cat!r}  history: {dict(hist[cat].most_common())}")
            for r in rs:
                payee = str(r[5] or "")[:30]
                pv = dict(hist_payee[cat].get(payee, Counter()).most_common(3))
                print(f"    tkey {r[0]} {r[3]} {r[2]:>9.2f}  {payee:<30} {str(r[6] or '')[:40]}")
                if pv:
                    print(f"        same payee before -> {pv}")
            out["ambiguous"][cat] = {"count": len(rs), "history": dict(hist[cat])}

    print("\n" + "-" * 70)
    print(f"New Categories currently available: {len(cats)}")
    print("  " + ", ".join(sorted(repr(c) for c in cats)))

    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=1, default=str)
        print(f"\nwritten -> {args.json}")


if __name__ == "__main__":
    main()
